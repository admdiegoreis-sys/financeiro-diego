import json
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\diego\OneDrive\01 - Financeiro_Diego\Novo financeiro\Financeiro_Diego_Consolidado.xlsx")
APP_DIR = Path(__file__).parent / "app"
DATA_JSON = APP_DIR / "finance-data.json"
DATA_JS = APP_DIR / "finance-data.js"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def number(value, default=0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def date_iso(value):
    if value is None or value == "":
        return ""
    if hasattr(value, "date"):
        return value.date().isoformat()
    text = str(value).strip()
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    return ""


def norm(value):
    return clean(value).lower().replace("\n", " ")


def operation(value):
    text = norm(value)
    if any(token in text for token in ["resgate", "venda", "saque"]):
        return "venda"
    return "compra"


def income_type(value):
    text = norm(value)
    if "juros" in text or text == "jcp":
        return "jcp"
    if "amort" in text:
        return "amortização"
    if "rendimento" in text:
        return "rendimento"
    if "dividend" in text:
        return "dividendo"
    return "outro"


def asset_type(value):
    text = norm(value)
    if "fundo imobili" in text or text == "fii":
        return "fii"
    if "renda fixa" in text or "cdb" in text or "tesouro" in text:
        return "renda fixa"
    if "fundo" in text:
        return "fundo"
    if "ação" in text or "acoes" in text or "ações" in text or "stock" in text:
        return "ação"
    return "outro"


def find_header_row(ws, required):
    required_set = {item.lower() for item in required}
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        labels = [norm(cell) for cell in row]
        if required_set.issubset(set(labels)):
            return labels
    return None


def row_dict(labels, row):
    return {label: row[index] for index, label in enumerate(labels) if label}


def parse_investment_summary(wb):
    summary = {}
    ws = wb["Investimentos"]
    labels = None
    for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
        row_labels = [norm(cell) for cell in row]
        if {"ativo", "nome", "tipo"}.issubset(set(row_labels)):
            labels = row_labels
            continue
        if labels and clean(row[1] if len(row) > 1 else ""):
            data = row_dict(labels, row)
            ticker = clean(data.get("ativo")).upper()
            if ticker:
                summary[ticker] = {
                    "name": clean(data.get("nome")),
                    "type": asset_type(data.get("tipo")),
                }
    return summary


def parse_investments(wb):
    summary = parse_investment_summary(wb)
    investments = []
    for ws in wb.worksheets:
        labels = find_header_row(ws, ["ativo", "corretora", "movimento", "data", "preço", "quant."])
        if not labels:
            continue
        header_index = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
            if [norm(cell) for cell in row] == labels:
                header_index = idx
                break
        if not header_index:
            continue
        for row_number, row in enumerate(ws.iter_rows(min_row=header_index + 1, values_only=True), start=header_index + 1):
            data = row_dict(labels, row)
            ticker = clean(data.get("ativo")).upper()
            date = date_iso(data.get("data"))
            qty = number(data.get("quant."))
            price = number(data.get("preço"))
            if not ticker or not date or not qty or not price:
                continue
            info = summary.get(ticker, {})
            investments.append({
                "id": f"excel-investment-{ws.title}-{row_number}",
                "date": date,
                "operation": operation(data.get("movimento")),
                "assetType": info.get("type") or asset_type(ws.title),
                "ticker": ticker,
                "assetName": info.get("name") or ticker,
                "quantity": qty,
                "unitPrice": price,
                "fees": number(data.get("taxas")),
                "broker": clean(data.get("corretora")),
                "notes": clean(data.get("movimento")),
                "source": ws.title,
            })
    return investments


def parse_incomes(wb):
    ws = wb["Rendimentos"]
    labels = find_header_row(ws, ["data", "produto", "tipo", "valor", "quantidade", "saldo bruto"])
    if not labels:
        return []
    header_index = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        if [norm(cell) for cell in row] == labels:
            header_index = idx
            break
    incomes = []
    for row_number, row in enumerate(ws.iter_rows(min_row=header_index + 1, values_only=True), start=header_index + 1):
        data = row_dict(labels, row)
        ticker = clean(data.get("produto")).upper()
        date = date_iso(data.get("data"))
        amount = number(data.get("saldo bruto"))
        qty = number(data.get("quantidade"))
        if not amount and qty:
            amount = number(data.get("valor")) * qty
        if not ticker or not date or not amount:
            continue
        incomes.append({
            "id": f"excel-income-{row_number}",
            "date": date,
            "type": income_type(data.get("descrição")),
            "ticker": ticker,
            "amount": amount,
            "quantity": qty,
            "account": clean(data.get("instituição")),
            "notes": clean(data.get("descrição")),
            "assetType": asset_type(data.get("tipo")),
            "source": "Rendimentos",
        })
    return incomes


def write_data(data):
    DATA_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    DATA_JS.write_text("window.FINANCE_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def main():
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    data = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    data["investments"] = parse_investments(wb)
    data["incomes"] = parse_incomes(wb)
    write_data(data)
    print(json.dumps({
        "investments": len(data["investments"]),
        "incomes": len(data["incomes"]),
        "investment_sources": sorted({row["source"] for row in data["investments"]})[:80],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
